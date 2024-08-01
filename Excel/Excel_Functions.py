import openpyxl

class ExcelFunctions:
    def __init__(self):
        pass

    def getRowCount(file, path, sheet_name):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        return sheet.max_row
    
    def getColumnCount(file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_column
    
    def readData(file, path, sheet_name, row, column):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row, column=column).value
    
    def writeData(file, path, sheet_name, row, column, data):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=column).value = data
        workbook.save(path)